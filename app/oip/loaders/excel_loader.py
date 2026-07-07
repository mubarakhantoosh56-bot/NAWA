"""Excel loading utilities for local NAWA OIP ingestion."""

from __future__ import annotations

import asyncio
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class ExcelSheet:
    """In-memory representation of one Excel sheet."""

    name: str
    rows: list[tuple[Any, ...]]


class ExcelLoader:
    """Load .xlsx workbooks without external services or database writes."""

    def load(self, path: str | Path) -> list[ExcelSheet]:
        """Load all workbook sheets from a local .xlsx file."""
        workbook_path = Path(path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"Excel file not found: {workbook_path}")
        if workbook_path.suffix.lower() != ".xlsx":
            raise ValueError(f"Unsupported Excel extension: {workbook_path.suffix}")

        workbook = load_workbook(
            io.BytesIO(workbook_path.read_bytes()),
            read_only=True,
            data_only=True,
        )
        try:
            sheets: list[ExcelSheet] = []
            for worksheet in workbook.worksheets:
                rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
                sheets.append(ExcelSheet(name=worksheet.title, rows=rows))
            return sheets
        finally:
            workbook.close()

    async def load_async(self, path: str | Path) -> list[ExcelSheet]:
        """Async-compatible wrapper for loading a local workbook."""
        return await asyncio.to_thread(self.load, path)

