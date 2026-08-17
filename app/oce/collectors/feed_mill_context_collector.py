"""Local feed mill context collection for OCE."""

from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import _env_bool
from app.oce.models.evidence import Evidence
from app.oip.loaders.excel_loader import ExcelLoader
from app.oip.translators.feed_mill_inventory_translator import FeedMillInventoryTranslator

# Resolution is directory-glob based (shape-based rule: filename is never
# authoritative). There is exactly one feed mill workbook in the pilot data
# source today, so this also stays correct if it is ever renamed.
FEED_MILL_DIR = (
    Path("data_sources")
    / "jannat_al_firdaws"
    / "2026_06"
    / "feed_mill"
)
FEED_MATERIAL_TERMS = (
    "علف",
    "ذرة",
    "الصويا",
    "نخالة",
    "بريمكس",
    "مثيونين",
    "ملح",
    "corn",
    "soya",
    "soybean",
    "bran",
    "premix",
    "feed",
    "salt",
)


@dataclass(frozen=True)
class FeedMillWorkbookSummary:
    """Basic local summary of a feed mill workbook."""

    path: Path
    sheet_names: list[str]
    detected_columns: list[str]
    row_count: int
    feed_related_row_count: int


class FeedMillContextCollector:
    """Collect local feed mill workbook evidence for operational context."""

    def __init__(
        self,
        loader: ExcelLoader | None = None,
        inventory_translator: FeedMillInventoryTranslator | None = None,
    ) -> None:
        self.loader = loader or ExcelLoader()
        self.inventory_translator = inventory_translator or FeedMillInventoryTranslator()

    def collect_evidence(
        self,
        date_range: tuple[Any, Any],
    ) -> Evidence | None:
        """Return available feed mill inventory evidence when a workbook can be read."""
        workbook_path = self._resolve_workbook_path()
        if workbook_path is None:
            return None

        try:
            summary = self._summarize_workbook(workbook_path)
        except Exception:
            return None

        description = (
            "Feed mill inventory workbook is available and readable. "
            f"Sheets: {', '.join(summary.sheet_names)}. "
            f"Detected columns/material labels: {', '.join(summary.detected_columns[:12])}. "
            f"Non-empty row count: {summary.row_count}. "
            f"Feed/material-related rows detected: {summary.feed_related_row_count}."
        )
        return Evidence(
            source=summary.path.as_posix(),
            type="feed_mill_inventory",
            status="available",
            description=description,
            date_range=date_range,
        )

    def collect_raw_material_inventory_evidence(
        self,
        date_range: tuple[Any, Any],
    ) -> Evidence | None:
        """Return structured raw_material_inventory evidence when the approved
        feed mill inventory snapshot shape/block is present in the workbook.

        This is distinct from ``collect_evidence`` (descriptive, whole-workbook
        readability) and from any poultry hall feed_received/feed_consumed
        evidence - the feed mill and poultry halls are different process
        stages per the Founder Business Semantics Ruling - M3.
        """
        workbook_path = self._resolve_workbook_path()
        if workbook_path is None:
            return None

        try:
            sheets = self.loader.load(workbook_path)
            records = self.inventory_translator.translate(sheets, workbook_path)
        except Exception:
            return None

        if not records:
            return None

        materials_with_inventory = sum(
            1 for record in records if record.raw_material_inventory is not None
        )
        materials_with_coverage = sum(
            1 for record in records if record.source_reported_days_coverage is not None
        )
        description = (
            "Feed mill raw material inventory snapshot is available "
            f"({len(records)} materials detected; "
            f"{materials_with_inventory} with a reported inventory balance, "
            f"{materials_with_coverage} with a source-reported days-of-coverage figure). "
            "No quantities are included in this description."
        )
        # M4 Slice 1 Golden Case: this evidence aggregates MULTIPLE material
        # records, so per-material fields (canonical_field/source_label/
        # raw_source_value/source_row_number) are deliberately left
        # unpopulated here rather than fabricated from one arbitrary
        # material (Founder instruction: do not pretend a multi-claim
        # artifact has one direct source claim). What IS uniform across
        # every record in one resolved block - and therefore safe to
        # surface - is the entity, the epistemic origin, the source
        # location, and critically the source's own report/snapshot time
        # status. ``date_range`` above remains the situation/reasoning
        # window; ``source_time``/``source_time_status`` below are the
        # SOURCE's own (possibly unresolved) date and must never be
        # conflated with it or silently backfilled from it.
        anchor = records[0]
        return Evidence(
            source=workbook_path.as_posix(),
            type="raw_material_inventory",
            status="available",
            description=description,
            date_range=date_range,
            epistemic_origin="observed",
            entity_type=anchor.entity_type,
            entity_reference=anchor.entity_reference,
            source_file=anchor.source_file,
            report_shape=anchor.report_shape,
            source_time=anchor.report_date,
            source_time_status=anchor.report_date_status,
            provenance_warnings=anchor.provenance_warnings,
        )

    def _resolve_workbook_path(self) -> Path | None:
        # M7 Slice 3A: the sole static-file read boundary for this
        # collector - both collect_evidence() and
        # collect_raw_material_inventory_evidence() call only this method
        # to find a workbook, so gating here protects the feed-mill static
        # scan regardless of which caller reaches it (including a future
        # caller other than PoultryContextCollector). Default True (unset
        # preserves today's behavior); a malformed value fails closed via
        # the existing _env_bool semantics (only recognized truthy strings
        # count as enabled). This is a legacy-STATIC-source switch only -
        # it never touches uploaded Truth, which has no static file to read.
        if not _env_bool("NAWA_STATIC_PILOT_DATA_SOURCES_ENABLED", True):
            return None
        if not FEED_MILL_DIR.exists():
            return None
        candidates = sorted(FEED_MILL_DIR.glob("*.xlsx"))
        return candidates[0] if candidates else None

    def _summarize_workbook(self, path: Path) -> FeedMillWorkbookSummary:
        workbook = load_workbook(io.BytesIO(path.read_bytes()), read_only=True, data_only=True)
        try:
            sheet_names = list(workbook.sheetnames)
            detected_columns: list[str] = []
            row_count = 0
            feed_related_row_count = 0
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(values_only=True):
                    values = [self._normalize_cell(value) for value in row]
                    non_empty = [value for value in values if value]
                    if not non_empty:
                        continue
                    row_count += 1
                    if self._looks_like_header(non_empty):
                        detected_columns.extend(non_empty)
                    if self._has_feed_material_term(non_empty):
                        feed_related_row_count += 1

            return FeedMillWorkbookSummary(
                path=path,
                sheet_names=sheet_names,
                detected_columns=self._dedupe(detected_columns),
                row_count=row_count,
                feed_related_row_count=feed_related_row_count,
            )
        finally:
            workbook.close()

    def _normalize_cell(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _looks_like_header(self, values: list[str]) -> bool:
        material_hits = sum(1 for value in values if self._has_feed_material_term([value]))
        return material_hits >= 2

    def _has_feed_material_term(self, values: list[str]) -> bool:
        text = " ".join(values).lower()
        return any(term.lower() in text for term in FEED_MATERIAL_TERMS)

    def _dedupe(self, values: list[str]) -> list[str]:
        seen: set[str] = set()
        unique: list[str] = []
        for value in values:
            if value in seen:
                continue
            seen.add(value)
            unique.append(value)
        return unique
